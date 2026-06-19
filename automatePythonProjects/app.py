import openpyxl as xl

# op
from openpyxl.chart import BarChart, Reference

# load_workbook is a function in the package.
wb = xl.load_workbook('transactions.xlsx')

# wb['Sheet1'] is used to access the particular sheet
sheet = wb['Sheet1']

cell = sheet['a1']
cell2 = sheet.cell(1, 1)
print(cell.value)
print(cell2)

# Reference class is used to select a range of values from xl sheet using rows and columns
values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
# chart is used to create chart
chart = BarChart()
# add_values() method of class Barchart is used to add values in the chart
chart.add_data(values)
# used to add
sheet.add_chart(chart, 'e2')

# Setting the header of the new column
new_cell_header = sheet.cell(1, 4)
new_cell_header.value = "Corrected Price Bruh"

# Looping from the 2nd row creating a new cell in each row and updating the price value
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)  # created a new cell in the same row
    corrected_price_cell.value = corrected_price

wb.save('transaction2.xlsx')
